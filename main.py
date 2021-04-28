#!/usr/bin/python3 

# Author: Molly Johnson
# Date: 3/24/21
# Description: Creates cboc signin excel file for
# checking in clinics for one month. Will adjust 
# for the days of the month, weekends, and federal holidays

####################################################################
### Function Title:
### Arguments:
### Returns:
### Description: 
####################################################################

# import openpyxl, datetime, and calendar
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import calendar
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

# create constants
JB = "Jesse Brown"
CP = "Crown Point"
HE = "Hoffman Est."
LA = "LaSalle"
AU = "Aurora"
JO = "Joliet"
KA = "Kankakee"
OL = "Oak Lawn"
FZ = "Frozen"
TCH = "Tech"
TOA = "Time of Arrival"
MID_DATE = 15
CBOC_COL_WIDTH = 12.5
HEADER_ROW_HEIGHT = 45
CBOC_ROW_HEIGHT = 20
DATE_ROW_HEIGHT = 22
TECH_TOA_ROW_HEIGHT = 27
CBOC_NAME_AND_FROZEN_ROW_HEIGHT = 21
SPACER_ROW_HEIGHT = 10
NUM_ROWS = 27
HEADER_ROW = 1
HEADER_AND_LABELS_COL = 1
CBOC_COL = 1
CBOC_ROW = 2
DATE_ROW = 3
TECH_TOA_ROW = 4
CBOC_NAME_AND_FROZEN_ROW_START = 5
CBOC_NAME_AND_FROZEN_ROWS = [5,6,8,9,11,12,14,15,17,18,20,21,23,24,26,27]
SPACER_ROWS = [7,10,13,16,19,22,25]
NAMES = [OL,HE,AU,KA,LA,JO,JB,CP] 
WEEKEND_AND_HOL_COL_WIDTH = 2.5
# set constant cell border values
thin = Side(border_style = "thin", color = "000000")
double = Side(border_style = "medium", color = "000000")
thick = Side(border_style = "thick", color = "001C54")
cbocNameBorder = Border(top = thick , left = thick, right = thick, bottom = thick) 
cbocNameFont = Font(name = 'Times New Roman', size = 10, bold = True)
dateBorder = Border(left = thick, right = thick, bottom = thick)
bigSpaceBorder = Border(left = thick, right = thick)
spacerBorder = Border(left = thick, right = thick)
cbocNameOnlyBorder = Border(top = double, left = thick, right = thick, bottom = thin)
frozenOnlyBorder = Border(left = thick, right = thick, bottom = double)
bottomRowBorderCBOCName = Border(left = thick, right = thick, bottom = thick)
dateFont = Font(name = 'Calibri', size = 11, bold = True)
techFont = Font(name = 'Calibri', size = 9)
toaFont = Font(name = 'Calibri', size = 5)
dateBorderLeft = Border(top = thick, left = thick, bottom = thick)
dateBorderRight = Border(top = thick, right = thick, bottom = thick)
techInfoBorderLeft = Border(left = thick, right = thin, bottom = double)
techInfoBorderRight = Border(right = thick, bottom = double)
sigBorderTopLeft = Border(top = double, left = thick, right = thin, bottom = thin)
sigBorderBottomLeft = Border(left = thick, right = thin, bottom = double)
sigBorderTopRight = Border(top = double, right = thick, bottom = thin)
sigBorderBottomRight = Border(right = thick, left = thin, bottom = double)
spacerBorderLeft = Border(right = thin, left = thick)
spacerBorderRight = Border(right = thick, left = thin)
weekendAndHolFillColor = PatternFill(fill_type = "solid", start_color = "BFBFBF", end_color = "BFBFBF")

####################################################################
### Function Title: setWeekendStyle()
### Arguments:
### Returns:
### Description: 
####################################################################
def setWeekendAndHolStyle(ws, curCol, curRow):
    endRow = NUM_ROWS
    spacerRows = [7,10,13,16,19,22,25]
    while(curRow <= endRow):
        ws.cell(row = curRow, column = curCol).fill = weekendAndHolFillColor      
        ws.cell(row = curRow, column = curCol + 1).fill = weekendAndHolFillColor      
        ws.column_dimensions[get_column_letter(curCol)].width = WEEKEND_AND_HOL_COL_WIDTH
        ws.column_dimensions[get_column_letter(curCol + 1)].width = WEEKEND_AND_HOL_COL_WIDTH
        if((curRow >= 5) and (curRow not in spacerRows)):
            ws.cell(row = curRow, column = curCol).value = "X"
            ws.cell(row = curRow, column = curCol + 1).value = "X"
            ws.cell(row = curRow, column = curCol).font = Font(name = 'Calibri', size = 15, bold = True)
            ws.cell(row = curRow, column = curCol + 1).font = Font(name = 'Calibri', size = 15, bold = True)
        curRow += 1

####################################################################
### Function Title: isWeekend()
### Arguments:
### Returns:
### Description: 
####################################################################
def isWeekend(dayName):
    if(dayName == "SAT" or dayName == "SUN"):
        return True
    return False

####################################################################
### Function Title: createSigBorders()
### Arguments:
### Returns:
### Description: 
####################################################################
def createSigBorders(ws, endCol):
    startCol = 2
    startRow = 5
    endRow = NUM_ROWS
    curCol = startCol
    endCbocNameRow = 26
    endSpacerRow = 25

    while (curCol <= endCol):
        # set cboc name row borders for both cols
        curRow = startRow
        while (curRow <= endCbocNameRow):
            # set left col border
            ws.cell(row = curRow, column = curCol).border = sigBorderTopLeft
            # set right col border
            ws.cell(row = curRow, column = curCol + 1).border = sigBorderTopRight
            curRow += 3

        # set frozen name row borders for both cols
        curRow = 6
        while (curRow <= endRow):
            # set left col border
            ws.cell(row = curRow, column = curCol).border = sigBorderBottomLeft
            # set right col border
            ws.cell(row = curRow, column = curCol + 1).border = sigBorderBottomRight
            if(curRow == endRow):
                # set left col border bottom to thick if last row
                ws.cell(row = curRow, column = curCol).border = Border(right = thin, bottom = thick)
                # set right col border bottom to thick if last row
                ws.cell(row = curRow, column = curCol + 1).border = Border(right = thick, bottom = thick)
            curRow += 3

        # set spacer row borders for both cols
        curRow = 7
        while (curRow <= endSpacerRow):
            # set left col border
            ws.cell(row = curRow, column = curCol).border = spacerBorderLeft
            # set right col border
            ws.cell(row = curRow, column = curCol + 1).border = spacerBorderRight
            curRow += 3    

        # increment cur col by 2 to move onto next date
        curCol += 2

####################################################################
### Function Title: mergeDate()
### Arguments:
### Returns:
### Description: 
####################################################################
def mergeDateInfo(ws, startRow, startCol):
    data = ws.cell(row = startRow, column = startCol).value
    ws.merge_cells(start_row = startRow, start_column = startCol, end_row = startRow, end_column = startCol + 1)
    ws.cell(row = startRow, column = startCol).value = data


####################################################################
### Function Title: setTechInfo()
### Arguments:
### Returns:
### Description: 
####################################################################
def setTechInfo(ws, curCol, curRow):
    ws.cell(row = curRow + 2, column = curCol).value = TCH
    ws.cell(row = curRow + 2, column = curCol).font = techFont
    ws.cell(row = curRow + 2, column = curCol).border = techInfoBorderLeft
    

    ws.cell(row = curRow + 2, column = curCol + 1).value = TOA
    ws.cell(row = curRow + 2, column = curCol + 1).font = toaFont
    ws.cell(row = curRow + 2, column = curCol + 1).border = techInfoBorderRight
    ws.cell(row = curRow + 2, column = curCol + 1).alignment = Alignment(wrap_text=True)

####################################################################
### Function Title: setDateInfo()
### Arguments:
### Returns:
### Description: 
####################################################################
def setDateInfo(ws, curCol, dayDate, dayName, dateNum, curRow):
    print('Day of Week (number): ', dayDate)
    print('Day of Week (name): ', dayName)
    print('Date of Month: ', dateNum)

    # set date name
    ws.cell(row = curRow, column = curCol).value = dayName
    mergeDateInfo(ws, curRow, curCol)
    ws.cell(row = curRow, column = curCol).font = dateFont
    ws.cell(row = curRow, column = curCol).alignment = Alignment(horizontal='center')
    ws.cell(row = curRow, column = curCol).border = dateBorderLeft
    ws.cell(row = curRow, column = curCol + 1).border = dateBorderRight

    # set date num
    ws.cell(row = curRow + 1, column = curCol).value = dateNum
    mergeDateInfo(ws, curRow + 1, curCol)
    ws.cell(row = curRow + 1, column = curCol).font = dateFont
    ws.cell(row = curRow + 1, column = curCol).alignment = Alignment(horizontal='center')
    ws.cell(row = curRow + 1, column = curCol).border = dateBorderLeft
    ws.cell(row = curRow + 1, column = curCol + 1).border = dateBorderRight

####################################################################
### Function Title: createDateCols()
### Arguments:
### Returns:
### Description: 
###################################################################
def createDateCols(ws, endCol, dateTimeObj, startDate, dayDate):
    #to get name of day (in number) from date
    # to get name of day from date
    dateNum = startDate
    dayName = calendar.day_abbr[dayDate]
    dayName = dayName.upper()

    # make start col and row 2 since dates start after CBOC col and header row
    curRow = 2
    curCol = 2
    #regColWidth = 3.67
    regColWidth = 4.5 
    while (curCol <= (endCol * 2)):
        
        ws.column_dimensions[get_column_letter(curCol)].width = regColWidth 

        # set the day date and name for the date/name cells
        setDateInfo(ws, curCol, dayDate, dayName, dateNum, curRow)
        #set tech info (tech, arrival time)
        setTechInfo(ws, curCol, curRow)

        
        
        # increment to get to second part of each date column
        curCol += 1
        ws.column_dimensions[get_column_letter(curCol)].width = regColWidth

        # check if is weekend
        if(isWeekend(dayName) == True):
            setWeekendAndHolStyle(ws, curCol - 1, curRow)

        # increment to get to first column of new date
        curCol += 1
        # increment the day and date
        dayDate += 1
        dateNum += 1
        #if day number is > 6, i.e. you've reached end of week, start week days over
        if (dayDate > 6):
            dayDate = 0
        dayName = calendar.day_abbr[dayDate]
        dayName = dayName.upper()

    return dayDate
        

####################################################################
### Function Title: createCBOCCOL()
### Arguments:
### Returns:
### Description: 
###################################################################
def createCBOCCol(ws):
    # create cboc col border and font
    
    # set cboc col width
    ws.column_dimensions[get_column_letter(CBOC_COL)].width = CBOC_COL_WIDTH
    
    # set cboc and date border and font
    ws.cell(row=2, column=CBOC_COL).font = cbocNameFont
    ws.cell(row=2, column=CBOC_COL).border = cbocNameBorder
    ws.cell(row=2, column=CBOC_COL).value = "CBOC/CORE"
    ws.cell(row=3, column=CBOC_COL).border = dateBorder
    ws.cell(row=3, column=CBOC_COL).font = cbocNameFont
    ws.cell(row=3, column=CBOC_COL).value = "Date"
    ws.cell(row=4, column=CBOC_COL).border = bigSpaceBorder

    # put in border/font for cboc name only rows
    i = 5
    j = 0
    while (i <= 26):
        ws.cell(row = i, column = CBOC_COL).font = cbocNameFont
        ws.cell(row = i, column = CBOC_COL).border = cbocNameOnlyBorder
        ws.cell(row = i, column = CBOC_COL).value = NAMES[j]
        i += 3
        j += 1

    # put in frozen rows
    i = 6
    while (i <= 27):
        ws.cell(row = i, column = CBOC_COL).font = cbocNameFont
        ws.cell(row = i, column = CBOC_COL).border = frozenOnlyBorder
        ws.cell(row = i, column = CBOC_COL).value = FZ
        if(i == 27):
            ws.cell(row = i, column = CBOC_COL).border = bottomRowBorderCBOCName
        i += 3

    # put in spacer rows
    i = 7
    while (i <= 25):
        ws.cell(row = i, column = CBOC_COL).border = spacerBorder
        i += 3
    
####################################################################
### Function Title: setRowHeights()
### Arguments:
### Returns:
### Description: 
####################################################################
def setRowHeights(ws):
    ws.row_dimensions[HEADER_ROW].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[CBOC_ROW].height = CBOC_ROW_HEIGHT 
    ws.row_dimensions[DATE_ROW].height = DATE_ROW_HEIGHT
    ws.row_dimensions[TECH_TOA_ROW].height = TECH_TOA_ROW_HEIGHT
    
    for rowNum in CBOC_NAME_AND_FROZEN_ROWS:
        ws.row_dimensions[rowNum].height = CBOC_NAME_AND_FROZEN_ROW_HEIGHT

    for rowNum in SPACER_ROWS:
        ws.row_dimensions[rowNum].height = SPACER_ROW_HEIGHT

####################################################################
### Function Title: validUserInput()
### Arguments:
### Returns:
### Description: 
####################################################################
def validUserInput(userInput):
    # check that length of user input string is correct
    if(len(userInput) != 5):
        return False
    
    # check that first two chars are digits, mid char
    # is / or -, and last 2 chars are digits.
    j = 0
    while(j < len(userInput)):
        if(j != 2):
            if(userInput[j].isdigit() == False):
                return False
        else:
            if(userInput[j] != '-' and userInput[j] != '/'):
                return False
        j += 1
    
    # if the month and year chars were digits, make sure they make sense
    monthInput = ""
    yearInput = ""
    monthInput += userInput[0]
    monthInput += userInput[1]
    yearInput += userInput[3]
    yearInput += userInput[4]

    # check that month is between 1 and 12
    if(int(monthInput) < 1 or int(monthInput) > 12):
        return False
    # check that year is between 2021 and 2099
    if (int(yearInput) < 21 or int(yearInput) > 99):
        return False
    
    # otherwise met all requirements, return true
    return True

####################################################################
### Function Title: getStartDate()
### Arguments:
### Returns:
### Description: 
####################################################################
def getStartDate():
    # get start date of the month from user
    i = 0
    while(i == 0):
        userInput = input("\nEnter month and year in the format mm/yy: ")
        if(validUserInput(userInput) == True):
            i = 1
        else:
            print("Your entry was invalid. Enter month and year in the format mm/yy or mm-yy:")

    # reformat start date input into string for datetime in format mm-01-20yy
    i = 0
    startDate = "" 
    while(i < len(userInput)):
        if(i >= 0 and i <= 1):
            startDate += userInput[i]
        elif(i == 2):
            startDate += "-01-20"
        elif(i >= 3):
            startDate += userInput[i]
        i += 1
    return startDate

####################################################################
### Function Title: createHeader()
### Arguments:
### Returns:
### Description: 
####################################################################
def createHeader(ws, startRow, startCol, endRow, endCol, startDateObj):
    # create header border formatting
    headerBorderLeft = Border(top = thick , left = thick, right = None, bottom = thick) 
    headerBorderRight = Border(top = thick , left = None, right = thick, bottom = thick)  
    headerBorderMid = Border(top = thick, left = None, right = None, bottom = thick)
    
    # font values
    headerFont = Font(name = 'Times New Roman', size = 28, bold = True)    
    
    ###################NEED TO USE NUMBERS NOT LETTERS FOR CELLS HERE
    #set header alignment to center, font to Times New Roman and size to 28
    #ws['A1'].alignment = Alignment(vertical = 'bottom')
    ws.cell(row = startRow, column = startCol).alignment = Alignment(vertical = 'bottom')
    #ws['A1'].alignment = Alignment(horizontal = 'center')
    ws.cell(row = startRow, column = startCol).alignment = Alignment(horizontal = 'center')
    ws.cell(row = startRow, column = startCol).font = headerFont
    
    # set border at far left and far right of header merged cells
    ws.cell(row = startRow, column = startCol).border = headerBorderLeft
    ws.cell(row = endRow, column = endCol).border = headerBorderRight
    
    # set border at middle header merged cells
    for row in ws.iter_rows(min_row = startRow, max_row = endRow, min_col = (startCol + 1), max_col = (endCol - 1)):
        for cell in row:
            cell.border = headerBorderMid 
            
    # create header and merge cells A1 through AE1
    #ws['A1'] = ("Month/Year: " + (str(calendar.month_name[startDateObj.month]) + " " + 
        #str(startDateObj.year)).upper())
    #data = ws['A1'].value 
    #ws.merge_cells('A1:AE1')
    #ws['A1'] = data 
    ws.cell(row = startRow, column = startCol).value =  ("Month/Year: " + (str(calendar.month_name[startDateObj.month]) + " " + str(startDateObj.year)).upper())
    data = ws.cell(row = startRow, column = startCol).value
    ws.merge_cells(start_row = startRow, start_column = startCol, end_row = endRow, end_column = endCol)
    ws.cell(row = startRow, column = startCol).value = data

####################################################################
### Function Title: getDateTimeObj()
### Arguments:
### Returns:
### Description: 
####################################################################   
def getDatetimeObj(startDate):
    dateTimeObj = datetime.strptime(startDate, "%m-%d-%Y")
    #print start date string
    print(startDate)
    #get day number from date
    print('Day of Month: ', dateTimeObj.day)
    #get year from date
    print('Year: ', dateTimeObj.year)
    #to get name of day (in number) from date
    print('Datetime day of Week (number): ', dateTimeObj.weekday())
    # to get name of day from date
    print('Datetime day of Week (name): ', calendar.day_abbr[dateTimeObj.weekday()])
    # to get name of month from date
    print('Month name: ', calendar.month_name[dateTimeObj.month])
    return dateTimeObj
    
####################################################################
### Function Title: main()
### Arguments:
### Returns:
### Description: 
####################################################################
def main():
    # create workbook (1st sheet at pos 0 created automatically)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "1-15"

    # create 2nd sheet at pos 1
    ws2 = wb.create_sheet("16-End", 1)
    
    # get start date string from user
    startDate = getStartDate()

    # Create date object in format mm-dd-yyyy from start date string
    startDateObj = getDatetimeObj(startDate)
    
    # get end date for the month
    endDate = calendar.monthrange(startDateObj.year, startDateObj.month)[1]
    print("End of month date: " + str(endDate))

    #to iterate to next date/day name
    #print('Next date (num) of week: ', (startDateObj.day + 1))
    #print('Next day of week (name): ', calendar.day_abbr[(startDateObj.weekday()) + 1])

    # set row height for all rows
    setRowHeights(ws1)
    setRowHeights(ws2)
    

    # create cboc cell font/border/values for both sheets
    createCBOCCol(ws1)
    createCBOCCol(ws2)
    
    # create rest of cols (date cols) for both sheets
    dayDate = startDateObj.weekday()
    # (update day date to be last day date from first sheet before passing to second sheet)
    dayDate = createDateCols(ws1, MID_DATE, startDateObj, 1, dayDate)
    createDateCols(ws2, endDate - MID_DATE, startDateObj, 16, dayDate)

    # create rest of borders for blank areas that will get signatures/initials and times
    createSigBorders(ws1, MID_DATE * 2)
    createSigBorders(ws2, (endDate - MID_DATE) * 2)
    
    # def createHeader(ws, startRow, startCol, endRow, endCol, startDateObj):
    # create header for both sheets
    createHeader(ws1, HEADER_ROW, HEADER_AND_LABELS_COL, HEADER_ROW, (MID_DATE * 2) + 1, startDateObj)
    createHeader(ws2, HEADER_ROW, HEADER_AND_LABELS_COL, HEADER_ROW, ((endDate - MID_DATE) * 2) + 1, startDateObj)
    print("end date minus mid date is: " + str(endDate) + " - " + str(MID_DATE) + " = " + str((endDate - MID_DATE)))
    
    # save workbook to excel file and exit
    wb.save('cboc_signin_sheet.xlsx')   

if __name__ == "__main__":
    main()